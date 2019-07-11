from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# 读取特定行，读取单双跳跃的那种，形成新的表格
# 预留一个费用日期，费用名称单数行，费用数量，费用单价，费用金额，费用类别，费用科室

excel = load_workbook("/Users/chenzuo/Desktop/test1.xlsx")

print(excel.sheetnames)
# 获取sheet：
table = excel["Sheet1"]  # 通过表名获取
# table = excel.get_sheet_by_name("Sheet1")

# 获取行数和列数：
rows = table.max_row  # 获取行数 从上往下数
cols = table.max_column  # 获取列数

# 写表格
workbook = Workbook()
booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet

fee_date = "1"

for i in range(rows):

    if i % 2 == 1 and i >= 11:
        # 9 11 13 15 16

        # 判断有没有，没有就沿用之前的
        # fee_date = "1"
        fee_date_1 = table.cell(row=i, column=1).value
        if fee_date_1 is not None:
            fee_date = fee_date_1

        # 费用名称
        fee_name = table.cell(row=i, column=5).value
        # 费用数量
        fee_num = table.cell(row=i, column=9).value
        # 费用单价
        fee_unit_price = table.cell(row=i, column=11).value
        # 费用金额
        fee_total_price = table.cell(row=i, column=13).value

        # 费用类别（需要特殊对待上下不同,如果当前行没有那么就在上一行）
        type_num = i;
        fee_type = table.cell(row=type_num, column=15).value
        if fee_type is None:
            type_num = i - 1
            fee_type = table.cell(row=type_num, column=15).value

        fee_type = table.cell(row=type_num, column=15).value
        # 费用科室
        fee_department = table.cell(row=i, column=17).value
        print(fee_name, fee_num, fee_unit_price, fee_total_price, fee_type, fee_department, i)

        list = [fee_date, fee_name, fee_num, fee_unit_price, fee_total_price, fee_type, fee_department]
        booksheet.append(list)

# 填写表格

# column 竖行 row横行
# max_row = rows -11
# max_col = cols
#
# # 设置自适应宽度
# for column_index in range(1, max_col + 1):
# 	maxWidth = 0
# 	for row_index in range(1, max_row + 1):
# 		value = booksheet.cell(row=row_index, column=column_index).value
# 		if isinstance(value, int):
# 			value = str(value)  # 将中间int类型改变
# 		print (value,row_index,column_index)
# 		chineseNum = (len(value.encode('utf-8')) - len(value)) / 2
# 		width = len(value.encode('utf-8')) - chineseNum + 4  # +4 留下多余空间
# 		if width > maxWidth:
# 			maxWidth = width
# 	codeAsc = ord(str(column_index)) + 16
# 	booksheet.column_dimensions[chr(codeAsc)].width = maxWidth
# # 这边要留着一个最大值
#
# #  居中显示
for meta_cell in booksheet:
    for cell in meta_cell:
        cell.alignment = Alignment(horizontal='center')

workbook.save("表格" + ".xlsx")








