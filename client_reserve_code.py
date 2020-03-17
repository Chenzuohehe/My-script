from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# 
# 
# 读取后六位条码 保存数据，读取列表2 写入数据

# 导出的数据excel （商品毛利分布表）
export_excel_path = "/Users/chenzuo/Desktop/code_wmjs.xlsx"
# 需要导入的excel （库存表）
import_excel_path = "/Users/chenzuo/Desktop/reserve.xlsx"
barcodeList = []

# 除去空格，获取条码
def arrange_string(string):
    if string is None:
        return ""
    string = str(string)
    string = string.replace(" ","")
    # string = string[-6:]
    return string

# 选择导出的数据
export_excel = load_workbook(export_excel_path)


# 获取sheet：
export_table = export_excel['商品毛利分布表']  # 通过表名获取
# export_table = export_excel.get_sheet_by_name("Sheet1")
# 获取行数和列数：
export_rows = export_table.max_row  # 获取行数 从上往下数

# 获取对应条码
for i in range(export_rows):
    if i % 2 == 0 and i > 0 and i < export_rows:
        # print(i)
        barcode = export_table.cell(row= i, column = 4).value
        # print (arrange_string(barcode))
        barcodeList.append(arrange_string(barcode))
        print(barcode)

# print (barcodeList)

print ("全条码个数",len(barcodeList))

workbook = Workbook()
booksheet = workbook.active 

import_excel = load_workbook(import_excel_path)
import_table = import_excel["即时库存表"]

# 获取行数和列数：
import_rows = import_table.max_row  # 获取行数 从上往下数
# 如果有对应条码那么在6的内容改为V
# print(import_rows)

for i in range(import_rows):
    # 获取的库存表的范围
    if i > 1 and i < 109:
        # 12
        # print (i)
        import_barcode = arrange_string(import_table.cell(row = i, column = 4).value)
        print(import_barcode)
        # 如果包含这个条码
        if import_barcode in barcodeList:
            # ddd = import_table.cell(row=i, column=6).value
            import_table.cell(row=i, column=10).value = "1"
            barcodeList.remove(import_barcode)

print ("没有在表格中")
print (barcodeList)
import_excel.save("表格" + ".xlsx")






