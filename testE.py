from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# 读取特定行，读取单双跳跃的那种，形成新的表格
# 预留一个费用日期，费用名称单数行，费用数量，费用单价，费用金额，费用类别，费用科室

excel = load_workbook("/Users/chenzuo/Desktop/321.xlsx")

print(excel.sheetnames)

# 获取sheet：
# table = excel['Sheet1']  # 通过表名获取
# table = excel.get_sheet_by_name("Sheet1")
# print(table)